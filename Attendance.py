import cv2
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
import face_recognition
import numpy as np

# === PATH SETUP ===
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FACES_DIR = os.path.join(BASE_DIR, "faces")
CASCADE_PATH = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")
EXCEL_FILE = os.path.join(BASE_DIR, "data.xlsx")
RECOGNIZED_FILE = os.path.join(BASE_DIR, "recognized_name.txt")

os.makedirs(FACES_DIR, exist_ok=True)
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Timestamp", "Image Path"])
    wb.save(EXCEL_FILE)

face_cascade = cv2.CascadeClassifier(CASCADE_PATH)

# === MAIN APP CLASS ===
class FaceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Attendance App")
        self.root.configure(bg="#228B22")

        self.name_entry = tk.Entry(root, font=("Arial", 16), bg="#B22222", fg="white", insertbackground="white")
        self.name_entry.pack(pady=10)

        self.video_label = tk.Label(root, bg="#228B22")
        self.video_label.pack()

        btn_frame = tk.Frame(root, bg="#228B22")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Register", command=self.register_face,
                  font=("Arial", 12, "bold"), bg="blue", fg="white", width=10).grid(row=0, column=0, padx=10)

        tk.Button(btn_frame, text="Sign In", command=self.sign_in_face,
                  font=("Arial", 12, "bold"), bg="green", fg="white", width=10).grid(row=0, column=1, padx=10)

        self.cap = cv2.VideoCapture(0)
        self.update_video()

    def update_video(self):
        ret, frame = self.cap.read()
        if not ret:
            return
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb)
        imgtk = ImageTk.PhotoImage(image=img)
        self.video_label.imgtk = imgtk
        self.video_label.configure(image=imgtk)
        self.current_frame = frame.copy()
        self.root.after(10, self.update_video)

    def register_face(self):
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showerror("Error", "Enter a name.")
            return
        face_path = os.path.join(FACES_DIR, f"{name}.jpg")
        cv2.imwrite(face_path, self.current_frame)
        messagebox.showinfo("Success", f"Face registered as {name}")

    def sign_in_face(self):
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showerror("Error", "Enter your name to sign in.")
            return

        saved_path = os.path.join(FACES_DIR, f"{name}.jpg")
        if not os.path.exists(saved_path):
            messagebox.showerror("Error", f"No face registered as '{name}'")
            return

        try:
            # === Load saved image safely ===
            pil_image = Image.open(saved_path).convert("RGB")
            known_image = np.array(pil_image)
            if known_image.dtype != np.uint8 or known_image.ndim != 3 or known_image.shape[2] != 3:
                raise ValueError("Invalid image format")

            known_encodings = face_recognition.face_encodings(known_image)
            if not known_encodings:
                raise ValueError("No face found in saved image")

            # === Get encoding from live webcam ===
            rgb_live = cv2.cvtColor(self.current_frame, cv2.COLOR_BGR2RGB)
            live_encodings = face_recognition.face_encodings(rgb_live)
            if not live_encodings:
                raise ValueError("No face detected in live image")

            match = face_recognition.compare_faces([known_encodings[0]], live_encodings[0])[0]
            if not match:
                messagebox.showerror("Error", "Face does not match saved image.")
                return

        except Exception as e:
            messagebox.showerror("Error", f"Face match failed:\n{str(e)}")
            return

        # === Log success to Excel and TXT ===
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, timestamp, saved_path])
        wb.save(EXCEL_FILE)

        with open(RECOGNIZED_FILE, "w") as f:
            f.write(name)

        messagebox.showinfo("Success", f"{name} signed in at {timestamp}")

# === RUN ===
if __name__ == "__main__":
    root = tk.Tk()
    app = FaceApp(root)
    root.mainloop()
