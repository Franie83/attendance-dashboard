import cv2
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Paths
IMAGE_DIR = "faces"
EXCEL_FILE = "data.xlsx"
TEXT_FILE = "Attendance list.txt"
RECOGNIZED_NAME_FILE = "recognized_name.txt"
HAAR_CASCADE_PATH = "haarcascade_frontalface_default.xml"  # Local XML file

# Ensure folders/files exist
os.makedirs(IMAGE_DIR, exist_ok=True)
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Timestamp", "Image Path"])
    wb.save(EXCEL_FILE)

# Load Haar cascade face detector from local file
face_cascade = cv2.CascadeClassifier(HAAR_CASCADE_PATH)

def generate_unique_filename(name):
    base = os.path.join(IMAGE_DIR, f"{name}.jpg")
    if not os.path.exists(base):
        return base
    counter = 1
    while True:
        new_name = os.path.join(IMAGE_DIR, f"{name}_{counter}.jpg")
        if not os.path.exists(new_name):
            return new_name
        counter += 1

class FaceCaptureApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WebCam")
        self.root.configure(bg="#228B22")  # Green background

        self.video_label = tk.Label(root, bg="#228B22")
        self.video_label.pack()

        self.name_entry = tk.Entry(root, font=("Arial", 14), bg="#B22222", fg="white", insertbackground="white")
        self.name_entry.pack(pady=10)

        self.capture_btn = tk.Button(
            root, text="Sign In",
            command=self.capture_and_submit,
            font=("Arial", 12, "bold"),
            bg="#B22222", fg="white",
            activebackground="#8B0000", activeforeground="white"
        )
        self.capture_btn.pack(pady=10)

        self.cap = cv2.VideoCapture(0)
        self.current_frame = None
        self.current_faces = []

        self.update_video()

    def update_video(self):
        ret, frame = self.cap.read()
        if not ret:
            return

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        self.current_frame = frame.copy()
        self.current_faces = faces

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb)
        imgtk = ImageTk.PhotoImage(image=img)
        self.video_label.imgtk = imgtk
        self.video_label.configure(image=imgtk)

        self.root.after(10, self.update_video)

    def capture_and_submit(self):
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showerror("Error", "Please enter a name.")
            return

        if os.path.exists(TEXT_FILE):
            with open(TEXT_FILE, "r") as f:
                signed_names = [line.strip().lower() for line in f.readlines()]
                if name.lower() in signed_names:
                    messagebox.showerror("Error", "You have already signed in.")
                    self.name_entry.delete(0, tk.END)
                    return

        if not hasattr(self, 'current_faces') or len(self.current_faces) == 0:
            messagebox.showerror("Error", "No face detected.")
            return

        if len(self.current_faces) > 1:
            messagebox.showwarning("Warning", "Only one face is allowed.")
            return

        image_path = generate_unique_filename(name)
        cv2.imwrite(image_path, self.current_frame)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, timestamp, image_path])
        wb.save(EXCEL_FILE)

        with open(TEXT_FILE, "a") as f:
            f.write(name + "\n")

        with open(RECOGNIZED_NAME_FILE, "w") as f:
            f.write(name)

        messagebox.showinfo("Success", "Completed.")

        self.cap.release()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = FaceCaptureApp(root)
    root.mainloop()
